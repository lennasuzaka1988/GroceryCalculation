import itertools
import urllib
import urllib.request

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl_image_loader import SheetImageLoader

from grocery_scraper import *
import import_grocery_list

import time

# Have to separate Sprouts and Hy-Vee lists into different spreadsheets as scraping only applies to Sprouts
# (Hy-Vee disallows scraping), info pulled from import_grocery_list module
grocery_list_xlsx = '../Grocery List.xlsx'


def price_and_image_scraping():
    wb = openpyxl.load_workbook(grocery_list_xlsx)
    ws_sprouts = wb['Sprouts']
    rows = dataframe_to_rows(product_df, index=False, header=False)

    # Apply value to cell if no value is listed
    for row_index, row in enumerate(rows, 1):
        for column_index, value in enumerate(row, 3):
            if ws_sprouts.cell(row=row_index, column=column_index).value and column_index == 3:
                pass
            else:
                ws_sprouts.cell(row=row_index, column=column_index, value=value)
    wb.save(grocery_list_xlsx)


def attach_excel_images():
    workbk = openpyxl.load_workbook(grocery_list_xlsx)
    ws_sprouts = workbk['Sprouts']
    image_loader = SheetImageLoader(ws_sprouts)

    img_urls = []
    cell_coordinates = []

    for image_row in ws_sprouts.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in image_row:
            img_urls.append(str(cell.value))
            cell_coordinates.append('F' + str(cell.row))

    file_names = []

    count = 1
    for i in img_urls:
        urllib.request.urlretrieve(i, f'image_0{str(count)}.png')
        file_names.append(f'image_0{str(count)}.png')
        count += 1

    file_dict = dict(zip(file_names, cell_coordinates))

    for file, location in file_dict.items():
        if image_loader.image_in(location):
            print('Image already exists!')
        else:
            img = Image(file)
            img.height = 65
            img.width = 65
            img.anchor = location
            ws_sprouts.add_image(img)
    workbk.save(grocery_list_xlsx)


# Searching for first product (requires separate function because of changed HTML in first product's results page) and
# subsequent products
def first_product_search():
    product = import_grocery_list.df_sprouts.iat[0, 1]
    first_product_result = product_info_list_output(product)
    all_prices_and_images.append(list(itertools.chain(*first_product_result)))
    # price_and_image_scraping()
    # attach_excel_images()
    # return first_product_result


def following_product_searches():
    for row in import_grocery_list.df_sprouts['Item'].dropna()[1:]:
        list_2_results = list(following_products_info_list_output(row))
        print(list_2_results)
        all_prices_and_images.append(list(itertools.chain(*list_2_results)))
        # price_and_image_scraping()
    # attach_excel_images()


# FIXME: Find out why subsequent searches aren't being performed properly after the first one
all_prices_and_images = []
store_navigation('64154')
time.sleep(5)
first_product_search()
time.sleep(5)
following_product_searches()
time.sleep(5)
# product_df = pd.DataFrame(all_prices_and_images, columns=['Price', 'Image Link'])
# price_and_image_scraping()



# TODO: Need to append product column with official name of product sought by scraper
# TODO: Develop way to wipe out price and image columns after completion of scraping and searching
